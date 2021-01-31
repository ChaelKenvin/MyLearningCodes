import openpyxl
from openpyxl .styles import Font

import os

# 合并和拆分单元格

os.chdir("D:\\Coding\\Python\\CH-12")

wb = openpyxl.Workbook()
sheet = wb[wb.sheetnames[0]]

sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells'

wb.save('merged.xlsx')

# 冻结窗格

# sheet.freeze_panes = 'A2'
sheet.freeze_panes = 'C1'

# A2    行1
# B1    列A
# C1    列A和列B
# A1 / none 没有冻结

wb.save('merged_freeze.xlsx')
