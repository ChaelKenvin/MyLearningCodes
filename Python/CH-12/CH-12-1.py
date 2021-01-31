import openpyxl
import os

os.chdir("D:\\Coding\\Python\\CH-12")

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))#表格的类型

print(wb.sheetnames)#获取表格子表格的名字

MySheet = wb['Shet1']#获取子表格
print(MySheet)
print(MySheet.title)

AnotherSheet = wb.sheetnames[0]#获取子表格
print(AnotherSheet)

print(MySheet['A1'])
print(MySheet['A1'].value)

#获取单元格
c = MySheet['B1']
print('coordinate: ' + c.coordinate)
print('B1 value: ' + str(c.value))
print('row: ' + str(c.row))
print('column: ' + str(c.column))

print(MySheet.cell(row = 1, column = 2))
print(MySheet.cell(1, 2))
print(MySheet.cell(1, 2).value)

for j in range(1, MySheet.max_row + 1):
    for i in range(1, MySheet.max_column + 1):
        print(MySheet.cell(j, i).value, end = ' ')
    print('\n')

print(MySheet.max_row)
print(MySheet.max_column)

