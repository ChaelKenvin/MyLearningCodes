import openpyxl
from openpyxl .styles import Font

import os

# 创建和保存excel文档

os.chdir("D:\\Coding\\Python\\CH-12")

wb = openpyxl.Workbook()  # 获取工作簿
print(wb.sheetnames)

sheet = wb[wb.sheetnames[0]]  # 获取工作表
sheet.title = 'Spam Bacon Egges sheet'  # 更改工作表的名字
print(wb.sheetnames)

wb.create_sheet()  # 在最后创建一个新工作表
print(wb.sheetnames)

wb.create_sheet(index=0, title='First Sheet')  # 在开头一个新工作表并命名
print(wb.sheetnames)

wb.create_sheet(index=1, title='Second Sheet')  # 在第二个位置插入一个新工作表并命名
print(wb.sheetnames)

wb.remove(wb['Second Sheet'])  # 移除工作表
print(wb.sheetnames)

sheet['A1'] = 'Hello World'  # 编辑单元格内容

italic24Font = Font(size=24, italic=True)
sheet['A1'].font = italic24Font
sheet['A1'] = 'Hello world!'

sheet = wb['Sheet']
styleObj1 = Font(size=16, name='Comic Sans MS', bold=True, italic=True)
sheet['A1'].font = styleObj1
sheet['A1'] = 'Hello World'

wb.save('example_copy.xlsx')  # 保存表格，这个表格在打开时不能保存

# wb2 = openpyxl.Workbook()
# sheet = wb2[wb2.sheetnames[0]]

# sheet['B1'] = 100
# sheet['B2'] = 200
# sheet['B3'] = 300
# sheet['B4'] = '=SUM(B1:B3)'
# print(sheet['B4'].value)

# wb2.save('example_copy2.xlsx')

# wbFornulas = openpyxl.load_workbook('example_copy2.xlsx')
# sheet = wbFornulas[wbFornulas.sheetnames[0]]
# print(sheet['B4'].value)

wbDataOnly = openpyxl.load_workbook('example_copy2.xlsx', data_only=True)
sheet = wbDataOnly[wbDataOnly.sheetnames[0]]
print(sheet['B4'].value)

